from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook


BASE_DIR = Path(__file__).resolve().parent.parent
EXPORT_DIR = BASE_DIR / "experiments"
RAW_LOG_PATH = EXPORT_DIR / "experimentation_runs.jsonl"
SUMMARY_CSV_PATH = EXPORT_DIR / "experimentation_summary.csv"
DETAILS_CSV_PATH = EXPORT_DIR / "experimentation_details.csv"
XLSX_PATH = EXPORT_DIR / "experimentation_results.xlsx"


SUMMARY_FIELDS = [
    "recorded_at",
    "test_name",
    "source",
    "lot_id",
    "site",
    "mineral_type",
    "http_status",
    "success",
    "result_status",
    "already_validated",
    "params_compared",
    "conformes",
    "non_conformes",
    "validated_by",
    "message",
    "error",
]

DETAIL_FIELDS = [
    "recorded_at",
    "test_name",
    "lot_id",
    "site",
    "result_status",
    "field",
    "label",
    "prodVal",
    "regVal",
    "diff",
    "tolerance",
    "ok",
]


def _ensure_export_dir() -> None:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def _to_bool_text(value: Any) -> str:
    return "oui" if bool(value) else "non"


def _normalize_record(record: dict[str, Any]) -> dict[str, Any]:
    comparison = record.get("comparison") or []
    conformes = int(record.get("conformes") or 0)
    params_compared = int(record.get("params_compared") or len(comparison))
    non_conformes = max(params_compared - conformes, 0)
    return {
        "recorded_at": record.get("recorded_at", ""),
        "test_name": record.get("test_name", "Auto-validation DGMR"),
        "source": record.get("source", "backend.auto_validate"),
        "lot_id": record.get("lot_id", ""),
        "site": record.get("site", ""),
        "mineral_type": record.get("mineral_type", ""),
        "http_status": record.get("http_status", ""),
        "success": _to_bool_text(record.get("success")),
        "result_status": record.get("result_status", ""),
        "already_validated": _to_bool_text(record.get("already_validated")),
        "params_compared": params_compared,
        "conformes": conformes,
        "non_conformes": non_conformes,
        "validated_by": record.get("validated_by", ""),
        "message": record.get("message", ""),
        "error": record.get("error", ""),
    }


def _load_records() -> list[dict[str, Any]]:
    if not RAW_LOG_PATH.exists():
        return []

    records: list[dict[str, Any]] = []
    with RAW_LOG_PATH.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            records.append(json.loads(line))
    return records


def _write_summary_csv(records: list[dict[str, Any]]) -> None:
    with SUMMARY_CSV_PATH.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=SUMMARY_FIELDS)
        writer.writeheader()
        for record in records:
            writer.writerow(_normalize_record(record))


def _write_details_csv(records: list[dict[str, Any]]) -> None:
    with DETAILS_CSV_PATH.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=DETAIL_FIELDS)
        writer.writeheader()
        for record in records:
            base = _normalize_record(record)
            comparison = record.get("comparison") or []
            if not comparison:
                writer.writerow({
                    "recorded_at": base["recorded_at"],
                    "test_name": base["test_name"],
                    "lot_id": base["lot_id"],
                    "site": base["site"],
                    "result_status": base["result_status"],
                    "field": "",
                    "label": "",
                    "prodVal": "",
                    "regVal": "",
                    "diff": "",
                    "tolerance": "",
                    "ok": "",
                })
                continue

            for item in comparison:
                writer.writerow({
                    "recorded_at": base["recorded_at"],
                    "test_name": base["test_name"],
                    "lot_id": base["lot_id"],
                    "site": base["site"],
                    "result_status": base["result_status"],
                    "field": item.get("field", ""),
                    "label": item.get("label", ""),
                    "prodVal": item.get("prodVal", ""),
                    "regVal": item.get("regVal", ""),
                    "diff": item.get("diff", ""),
                    "tolerance": item.get("tolerance", ""),
                    "ok": _to_bool_text(item.get("ok")),
                })


def _autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)


def _write_workbook(records: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Synthese"
    summary_ws.append(SUMMARY_FIELDS)

    for record in records:
        normalized = _normalize_record(record)
        summary_ws.append([normalized[field] for field in SUMMARY_FIELDS])

    details_ws = workbook.create_sheet("Comparaisons")
    details_ws.append(DETAIL_FIELDS)
    for record in records:
        base = _normalize_record(record)
        comparison = record.get("comparison") or []
        if not comparison:
            details_ws.append([
                base["recorded_at"],
                base["test_name"],
                base["lot_id"],
                base["site"],
                base["result_status"],
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ])
            continue

        for item in comparison:
            details_ws.append([
                base["recorded_at"],
                base["test_name"],
                base["lot_id"],
                base["site"],
                base["result_status"],
                item.get("field", ""),
                item.get("label", ""),
                item.get("prodVal", ""),
                item.get("regVal", ""),
                item.get("diff", ""),
                item.get("tolerance", ""),
                _to_bool_text(item.get("ok")),
            ])

    meta_ws = workbook.create_sheet("Meta")
    meta_ws.append(["champ", "valeur"])
    meta_ws.append(["exported_at", datetime.utcnow().isoformat()])
    meta_ws.append(["runs_count", len(records)])
    meta_ws.append(["summary_csv", str(SUMMARY_CSV_PATH.name)])
    meta_ws.append(["details_csv", str(DETAILS_CSV_PATH.name)])

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        _autosize_worksheet(ws)

    workbook.save(XLSX_PATH)


def rebuild_exports() -> dict[str, str]:
    _ensure_export_dir()
    records = _load_records()
    _write_summary_csv(records)
    _write_details_csv(records)
    _write_workbook(records)
    return {
        "raw_log": str(RAW_LOG_PATH),
        "summary_csv": str(SUMMARY_CSV_PATH),
        "details_csv": str(DETAILS_CSV_PATH),
        "xlsx": str(XLSX_PATH),
    }


def record_auto_validation_run(
    *,
    lot_id: str,
    http_status: int,
    success: bool,
    result_status: str = "",
    already_validated: bool = False,
    params_compared: int = 0,
    conformes: int = 0,
    validated_by: str = "",
    site: str = "",
    mineral_type: str = "",
    message: str = "",
    error: str = "",
    comparison: list[dict[str, Any]] | None = None,
    test_name: str = "Auto-validation DGMR",
    source: str = "backend.auto_validate",
) -> dict[str, str]:
    _ensure_export_dir()
    record = {
        "recorded_at": datetime.utcnow().isoformat(),
        "test_name": test_name,
        "source": source,
        "lot_id": lot_id,
        "site": site,
        "mineral_type": mineral_type,
        "http_status": http_status,
        "success": success,
        "result_status": result_status,
        "already_validated": already_validated,
        "params_compared": params_compared,
        "conformes": conformes,
        "validated_by": validated_by,
        "message": message,
        "error": error,
        "comparison": comparison or [],
    }
    with RAW_LOG_PATH.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(record, ensure_ascii=False) + "\n")
    return rebuild_exports()


def replace_records(records: list[dict[str, Any]]) -> dict[str, str]:
    _ensure_export_dir()
    with RAW_LOG_PATH.open("w", encoding="utf-8") as handle:
        for record in records:
            normalized = {
                "recorded_at": record.get("recorded_at", datetime.utcnow().isoformat()),
                "test_name": record.get("test_name", "Cas de test"),
                "source": record.get("source", "historique.documentation"),
                "lot_id": record.get("lot_id", ""),
                "site": record.get("site", ""),
                "mineral_type": record.get("mineral_type", ""),
                "http_status": record.get("http_status", ""),
                "success": record.get("success", True),
                "result_status": record.get("result_status", ""),
                "already_validated": record.get("already_validated", False),
                "params_compared": record.get("params_compared", 0),
                "conformes": record.get("conformes", 0),
                "validated_by": record.get("validated_by", ""),
                "message": record.get("message", ""),
                "error": record.get("error", ""),
                "comparison": record.get("comparison", []),
            }
            handle.write(json.dumps(normalized, ensure_ascii=False) + "\n")
    return rebuild_exports()
